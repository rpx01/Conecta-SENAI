"""API do módulo planejamento."""
from __future__ import annotations

import io
from datetime import datetime
from flask import jsonify, request, abort, send_file
from sqlalchemy import extract
from openpyxl import load_workbook, Workbook

from src.auth import require_roles, ROLE_ADMIN, ROLE_GESTOR, ROLE_USER
from src.models import db
from src.models.planejamento import Planejamento
from src.models.instrutor import Instrutor

from . import bp


def _get_or_create_instrutor(nome: str) -> Instrutor:
    instrutor = Instrutor.query.filter_by(nome=nome).first()
    if not instrutor:
        instrutor = Instrutor(nome=nome)
        db.session.add(instrutor)
        db.session.flush()
    return instrutor


@bp.get("/api/planejamento")
@require_roles(ROLE_ADMIN, ROLE_GESTOR, ROLE_USER)
def api_list():
    mes = request.args.get("mes")
    q = Planejamento.query
    if mes:
        y, m = mes.split("-")
        q = q.filter(
            extract("year", Planejamento.data) == int(y),
            extract("month", Planejamento.data) == int(m),
        )
    instrutor_id = request.args.get("instrutor_id")
    status = request.args.get("status")
    if instrutor_id:
        q = q.filter_by(instrutor_id=int(instrutor_id))
    if status:
        q = q.filter_by(status=status)
    itens = q.order_by(Planejamento.data.asc()).all()
    return jsonify(
        [
            {
                "id": p.id,
                "data": p.data.isoformat(),
                "turno": p.turno,
                "carga_horas": p.carga_horas,
                "modalidade": p.modalidade,
                "treinamento": p.treinamento,
                "instrutor_id": p.instrutor_id,
                "instrutor": p.instrutor.nome if p.instrutor else None,
                "local": p.local,
                "cliente": p.cliente,
                "observacao": p.observacao,
                "status": p.status,
            }
            for p in itens
        ]
    )


@bp.post("/api/planejamento")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_create():
    data = request.get_json()
    dt = datetime.strptime(data["data"], "%Y-%m-%d").date()
    exists = Planejamento.query.filter_by(
        data=dt, turno=data["turno"], instrutor_id=data["instrutor_id"]
    ).first()
    if exists:
        return jsonify({"error": "Conflito de agenda"}), 409
    p = Planejamento(
        data=dt,
        turno=data["turno"],
        carga_horas=data.get("carga_horas"),
        modalidade=data.get("modalidade"),
        treinamento=data["treinamento"],
        instrutor_id=data["instrutor_id"],
        local=data.get("local"),
        cliente=data.get("cliente"),
        observacao=data.get("observacao"),
        status=data.get("status", "Planejado"),
        origem=data.get("origem", "Manual"),
    )
    db.session.add(p)
    db.session.commit()
    return jsonify({"id": p.id}), 201


@bp.put("/api/planejamento/<int:pid>")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_update(pid: int):
    p = Planejamento.query.get_or_404(pid)
    data = request.get_json()
    for f in [
        "data",
        "turno",
        "carga_horas",
        "modalidade",
        "treinamento",
        "instrutor_id",
        "local",
        "cliente",
        "observacao",
        "status",
    ]:
        if f in data:
            if f == "data":
                setattr(p, f, datetime.strptime(data[f], "%Y-%m-%d").date())
            else:
                setattr(p, f, data[f])
    exists = Planejamento.query.filter_by(
        data=p.data, turno=p.turno, instrutor_id=p.instrutor_id
    ).first()
    if exists and exists.id != p.id:
        return jsonify({"error": "Conflito de agenda"}), 409
    db.session.commit()
    return jsonify({"ok": True})


@bp.delete("/api/planejamento/<int:pid>")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_delete(pid: int):
    p = Planejamento.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({"ok": True})


@bp.post("/api/planejamento/import")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_import():
    f = request.files["file"]
    wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
    ws = wb.active
    headers = [
        str((ws.cell(row=1, column=i).value or "")).strip().upper()
        for i in range(1, ws.max_column + 1)
    ]

    def col(name: str) -> int | None:
        return headers.index(name) + 1 if name in headers else None

    if (
        col("DATA")
        and col("TURNO")
        and col("TREINAMENTO")
        and col("INSTRUTOR")
    ):
        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, col("DATA")).value
            if not d:
                continue
            data_val = (
                d.date()
                if hasattr(d, "date")
                else datetime.strptime(str(d), "%d/%m/%Y").date()
            )
            turno = (ws.cell(r, col("TURNO")).value or "").upper()[:5]
            instrutor_nome = (
                str(ws.cell(r, col("INSTRUTOR")).value or "").strip()
            )
            instrutor = _get_or_create_instrutor(instrutor_nome)
            p = Planejamento(
                data=data_val,
                turno=turno,
                carga_horas=ws.cell(r, col("CARGA") or 0).value,
                modalidade=ws.cell(r, col("MODALIDADE") or 0).value,
                treinamento=str(ws.cell(r, col("TREINAMENTO")).value or ""),
                instrutor_id=instrutor.id,
                local=str(ws.cell(r, col("LOCAL") or 0).value or ""),
                cliente=str(ws.cell(r, col("CLIENTE") or 0).value or ""),
                observacao=str(ws.cell(r, col("OBS") or 0).value or ""),
            )
            db.session.add(p)
        db.session.commit()
        return jsonify({"ok": True})
    abort(400, "Formato de planilha não reconhecido")


@bp.get("/api/planejamento/export")
@require_roles(ROLE_ADMIN, ROLE_GESTOR, ROLE_USER)
def api_export():
    mes = request.args.get("mes")
    y, m = [int(x) for x in mes.split("-")]
    q = Planejamento.query.filter(
        extract("year", Planejamento.data) == y,
        extract("month", Planejamento.data) == m,
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Planejamento"
    ws.append(
        [
            "DATA",
            "SEMANA",
            "TURNO",
            "CARGA",
            "MODALIDADE",
            "TREINAMENTO",
            "INSTRUTOR",
            "LOCAL",
            "CLIENTE",
            "STATUS",
            "OBSERVACAO",
        ]
    )
    for p in q:
        d = p.data
        ws.append(
            [
                d.strftime("%d/%m/%Y"),
                ["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sab"][d.weekday()],
                p.turno,
                p.carga_horas,
                p.modalidade,
                p.treinamento,
                p.instrutor.nome if p.instrutor else "",
                p.local,
                p.cliente,
                p.status,
                p.observacao or "",
            ]
        )
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"planejamento_{mes}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
