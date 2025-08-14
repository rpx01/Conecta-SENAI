"""Endpoints de API para o módulo de planejamento."""
from datetime import date, datetime
import io

from flask import jsonify, request, abort, send_file
from sqlalchemy import extract
from openpyxl import load_workbook, Workbook

from app.auth import require_roles, ROLE_ADMIN, ROLE_GESTOR, ROLE_USER
from app.extensions import db
from app.planejamento import bp
from app.planejamento.models import Planejamento
from src.models.instrutor import Instrutor


def get_or_create_instrutor(nome: str) -> Instrutor:
    """Obtém um instrutor pelo nome ou cria um novo caso não exista."""
    instrutor = Instrutor.query.filter_by(nome=nome).first()
    if not instrutor:
        instrutor = Instrutor(nome=nome)
        db.session.add(instrutor)
        db.session.flush()
    return instrutor


@bp.get("/api/planejamento")
@require_roles(ROLE_ADMIN, ROLE_GESTOR, ROLE_USER)
def api_list():
    """Lista itens de planejamento com filtros simples."""
    mes = request.args.get("mes")  # YYYY-MM
    q = Planejamento.query
    if mes:
        try:
            y, m = mes.split("-")
            q = q.filter(
                extract("year", Planejamento.data) == int(y),
                extract("month", Planejamento.data) == int(m),
            )
        except ValueError:
            abort(400, "Formato de mês inválido. Use YYYY-MM")
    itens = q.order_by(Planejamento.data.asc()).all()
    return jsonify([
        {
            "id": p.id,
            "data": p.data.isoformat(),
            "turno": p.turno,
            "carga_horas": p.carga_horas,
            "modalidade": p.modalidade,
            "treinamento": p.treinamento,
            "instrutor_id": p.instrutor_id,
            "instrutor": p.instrutor.nome if p.instrutor else None,
            "local": p.local,
            "cliente": p.cliente,
            "observacao": p.observacao,
            "status": p.status,
        }
        for p in itens
    ])


@bp.post("/api/planejamento")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_create():
    """Cria um novo item de planejamento."""
    data = request.get_json() or {}
    p = Planejamento(
        data=data["data"],
        turno=data["turno"],
        carga_horas=data.get("carga_horas"),
        modalidade=data.get("modalidade"),
        treinamento=data["treinamento"],
        instrutor_id=data["instrutor_id"],
        local=data.get("local"),
        cliente=data.get("cliente"),
        observacao=data.get("observacao"),
        status=data.get("status", "Planejado"),
        origem=data.get("origem", "Manual"),
    )
    db.session.add(p)
    db.session.commit()
    return jsonify({"id": p.id}), 201


@bp.put("/api/planejamento/<int:pid>")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_update(pid: int):
    """Atualiza um item de planejamento existente."""
    p = Planejamento.query.get_or_404(pid)
    data = request.get_json() or {}
    for f in [
        "data",
        "turno",
        "carga_horas",
        "modalidade",
        "treinamento",
        "instrutor_id",
        "local",
        "cliente",
        "observacao",
        "status",
    ]:
        if f in data:
            setattr(p, f, data[f])
    db.session.commit()
    return jsonify({"ok": True})


@bp.delete("/api/planejamento/<int:pid>")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_delete(pid: int):
    """Remove um item de planejamento."""
    p = Planejamento.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({"ok": True})


@bp.post("/api/planejamento/import")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_import():
    # Aceitar arquivo .xlsx com 2 formatos:
    # (A) Lista com colunas: DATA, SEMANA, TURNO, CARGA, MODALIDADE, TREINAMENTO, INSTRUTOR, LOCAL, CLIENTE, OBS
    # (B) Matriz (Data x Instrutores): célula contém o nome do treinamento; turno vem de coluna dedicada
    f = request.files["file"]
    wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
    ws = wb.active
    headers = [
        str((ws.cell(row=1, column=i).value or "")).strip().upper()
        for i in range(1, ws.max_column + 1)
    ]

    def col(name):
        return headers.index(name) + 1 if name in headers else None

    if col("DATA") and col("TURNO") and col("TREINAMENTO") and col("INSTRUTOR"):
        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, col("DATA")).value
            if not d:
                continue
            data = d.date() if hasattr(d, "date") else datetime.strptime(str(d), "%d/%m/%Y").date()
            turno = (
                (ws.cell(r, col("TURNO")).value or "")
                .upper()[:5]
                .replace("Ã", "A")
                .replace("Â", "A")
            )
            instrutor_nome = str(ws.cell(r, col("INSTRUTOR")).value or "").strip()
            instrutor = get_or_create_instrutor(instrutor_nome)
            p = Planejamento(
                data=data,
                turno=turno,
                carga_horas=ws.cell(r, col("CARGA")).value if col("CARGA") else None,
                modalidade=ws.cell(r, col("MODALIDADE")).value if col("MODALIDADE") else None,
                treinamento=ws.cell(r, col("TREINAMENTO")).value,
                instrutor_id=instrutor.id,
                local=ws.cell(r, col("LOCAL")).value if col("LOCAL") else None,
                cliente=ws.cell(r, col("CLIENTE")).value if col("CLIENTE") else None,
                observacao=ws.cell(r, col("OBS")).value if col("OBS") else None,
                origem="Importado",
            )
            db.session.add(p)
        db.session.commit()
        return jsonify({"ok": True})
    # TODO: tratar formato B (matriz) — mapear cabeçalho como instrutores e varrer linhas por data/turno.
    abort(400, "Formato de planilha não reconhecido")


@bp.get("/api/planejamento/export")
@require_roles(ROLE_ADMIN, ROLE_GESTOR, ROLE_USER)
def api_export():
    mes = request.args.get("mes")
    y, m = [int(x) for x in mes.split("-")]
    q = Planejamento.query.filter(
        extract("year", Planejamento.data) == y,
        extract("month", Planejamento.data) == m,
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Planejamento"
    ws.append([
        "DATA",
        "SEMANA",
        "TURNO",
        "CARGA",
        "MODALIDADE",
        "TREINAMENTO",
        "INSTRUTOR",
        "LOCAL",
        "CLIENTE",
        "STATUS",
        "OBSERVACAO",
    ])
    for p in q:
        d = p.data
        ws.append([
            d.strftime("%d/%m/%Y"),
            ["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sab"][d.weekday()],
            p.turno,
            p.carga_horas,
            p.modalidade,
            p.treinamento,
            p.instrutor.nome if p.instrutor else "",
            p.local,
            p.cliente,
            p.status,
            p.observacao or "",
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"planejamento_{mes}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
