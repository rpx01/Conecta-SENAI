"""Rotas e API para o módulo de planejamento."""
import io
from datetime import date, datetime
from flask import (
    Blueprint,
    render_template,
    g,
    request,
    jsonify,
    abort,
    send_file,
)
from sqlalchemy import extract
from openpyxl import load_workbook, Workbook

from src.auth import login_required, admin_required
from src.models import db, Instrutor, Planejamento

planejamento_bp = Blueprint("planejamento", __name__, template_folder='../templates')


def get_or_create_instrutor(nome: str) -> Instrutor:
    """Obtém um instrutor pelo nome ou cria um novo caso não exista."""
    instrutor = Instrutor.query.filter_by(nome=nome).first()
    if not instrutor:
        instrutor = Instrutor(nome=nome)
        db.session.add(instrutor)
        db.session.flush()
    return instrutor


@planejamento_bp.app_context_processor
def inject_globals():
    """Injeta variáveis globais nos templates."""
    user_role = ""
    if hasattr(g, "current_user") and g.current_user:
        user_role = g.current_user.tipo.upper()
    return {
        "user_role": user_role,
        "current_year_month": date.today().strftime("%Y-%m"),
    }


# --- Rotas Web ---
@planejamento_bp.route("/")
@login_required
def index():
    """Página inicial do módulo de planejamento."""
    return render_template("planejamento/index.html")


@planejamento_bp.route("/matriz")
@login_required
def matriz():
    """Página com a matriz de planejamento."""
    return render_template("planejamento/matriz.html")


@planejamento_bp.route("/lista")
@login_required
def lista():
    """Página com a lista de itens planejados."""
    return render_template("planejamento/lista.html")


# --- API Endpoints ---
@planejamento_bp.get("/api/planejamento")
@login_required
def api_list():
    """Lista itens de planejamento com filtros."""
    mes = request.args.get("mes")
    q = Planejamento.query
    if mes:
        try:
            y, m = mes.split("-")
            q = q.filter(
                extract("year", Planejamento.data) == int(y),
                extract("month", Planejamento.data) == int(m),
            )
        except ValueError:
            abort(400, "Formato de mês inválido. Use YYYY-MM")
    itens = q.order_by(Planejamento.data.asc()).all()
    return jsonify([p.to_dict() for p in itens])


@planejamento_bp.post("/api/planejamento")
@admin_required
def api_create():
    """Cria um novo item de planejamento."""
    data = request.get_json() or {}
    exists = Planejamento.query.filter_by(
        data=data["data"],
        turno=data["turno"],
        instrutor_id=data["instrutor_id"],
    ).first()
    if exists:
        return jsonify({"error": "Conflito de agenda"}), 409
    p = Planejamento(
        data=datetime.strptime(data["data"], "%Y-%m-%d").date(),
        turno=data["turno"],
        carga_horas=data.get("carga_horas"),
        modalidade=data.get("modalidade"),
        treinamento=data["treinamento"],
        instrutor_id=data["instrutor_id"],
        local=data.get("local"),
        cliente=data.get("cliente"),
        observacao=data.get("observacao"),
        status=data.get("status", "Planejado"),
        origem="Manual",
    )
    db.session.add(p)
    db.session.commit()
    return jsonify({"id": p.id}), 201


@planejamento_bp.put("/api/planejamento/<int:pid>")
@admin_required
def api_update(pid: int):
    """Atualiza um item de planejamento existente."""
    p = Planejamento.query.get_or_404(pid)
    data = request.get_json() or {}
    new_data = data.get("data", p.data.isoformat())
    new_turno = data.get("turno", p.turno)
    new_instrutor = data.get("instrutor_id", p.instrutor_id)

    exists = Planejamento.query.filter(
        Planejamento.data == new_data,
        Planejamento.turno == new_turno,
        Planejamento.instrutor_id == new_instrutor,
        Planejamento.id != pid
    ).first()

    if exists:
        return jsonify({"error": "Conflito de agenda"}), 409

    for f in [
        "data", "turno", "carga_horas", "modalidade", "treinamento",
        "instrutor_id", "local", "cliente", "observacao", "status",
    ]:
        if f in data:
            if f == 'data':
                setattr(p, f, datetime.strptime(data[f], "%Y-%m-%d").date())
            else:
                setattr(p, f, data[f])

    db.session.commit()
    return jsonify({"ok": True})


@planejamento_bp.delete("/api/planejamento/<int:pid>")
@admin_required
def api_delete(pid: int):
    """Remove um item de planejamento."""
    p = Planejamento.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({"ok": True})


@planejamento_bp.post("/api/planejamento/import")
@admin_required
def api_import():
    """Importa dados de planejamento de uma planilha .xlsx."""
    f = request.files.get("file")
    if not f:
        abort(400, "Arquivo não enviado.")

    wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=i).value or "").strip().upper() for i in range(1, ws.max_column + 1)]

    def col(name):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    if all(col(c) for c in ["DATA", "TURNO", "TREINAMENTO", "INSTRUTOR"]):
        for r in range(2, ws.max_row + 1):
            d_val = ws.cell(r, col("DATA")).value
            if not d_val:
                continue

            data = d_val.date() if isinstance(d_val, datetime) else datetime.strptime(str(d_val), "%d/%m/%Y").date()
            turno = str(ws.cell(r, col("TURNO")).value or "").upper().strip()
            instrutor_nome = str(ws.cell(r, col("INSTRUTOR")).value or "").strip()
            instrutor = get_or_create_instrutor(instrutor_nome)

            p = Planejamento(
                data=data,
                turno=turno,
                treinamento=ws.cell(r, col("TREINAMENTO")).value,
                instrutor_id=instrutor.id,
                origem="Importado",
            )
            db.session.add(p)
        db.session.commit()
        return jsonify({"ok": True})

    abort(400, "Formato de planilha não reconhecido.")


@planejamento_bp.get("/api/planejamento/export")
@login_required
def api_export():
    """Exporta o planejamento do mês em formato .xlsx."""
    mes = request.args.get("mes")
    y, m = [int(x) for x in mes.split("-")]
    q = Planejamento.query.filter(
        extract("year", Planejamento.data) == y,
        extract("month", Planejamento.data) == m,
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Planejamento"
    ws.append(["DATA", "SEMANA", "TURNO", "INSTRUTOR", "TREINAMENTO", "STATUS"])
    for p in q:
        d = p.data
        ws.append([
            d.strftime("%d/%m/%Y"),
            d.strftime("%A"),
            p.turno,
            p.instrutor.nome if p.instrutor else "",
            p.treinamento,
            p.status,
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"planejamento_{mes}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
