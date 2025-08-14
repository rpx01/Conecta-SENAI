from datetime import date, datetime
import io
from flask import jsonify, request, abort, send_file
from sqlalchemy import extract
from openpyxl import load_workbook, Workbook

from src.auth import require_roles, ROLE_ADMIN, ROLE_GESTOR, ROLE_USER
from src.models import db
from src.models.instrutor import Instrutor
from .models import Planejamento
from . import bp


def _get_instrutor_id(nome: str) -> int:
    instr = Instrutor.query.filter_by(nome=nome).first()
    if not instr:
        instr = Instrutor(nome=nome)
        db.session.add(instr)
        db.session.flush()
    return instr.id


@bp.get("/api/planejamento")
@require_roles(ROLE_ADMIN, ROLE_GESTOR, ROLE_USER)
def api_list():
    mes = request.args.get("mes")
    instrutor_id = request.args.get("instrutor_id", type=int)
    status = request.args.get("status")
    q = Planejamento.query
    if mes:
        y, m = mes.split("-")
        q = q.filter(
            extract("year", Planejamento.data) == int(y),
            extract("month", Planejamento.data) == int(m),
        )
    if instrutor_id:
        q = q.filter(Planejamento.instrutor_id == instrutor_id)
    if status:
        q = q.filter(Planejamento.status == status)
    itens = q.order_by(Planejamento.data.asc()).all()
    return jsonify([
        {
            "id": p.id,
            "data": p.data.isoformat(),
            "turno": p.turno,
            "carga_horas": p.carga_horas,
            "modalidade": p.modalidade,
            "treinamento": p.treinamento,
            "instrutor_id": p.instrutor_id,
            "instrutor": p.instrutor.nome if p.instrutor else None,
            "local": p.local,
            "cliente": p.cliente,
            "observacao": p.observacao,
            "status": p.status,
        }
        for p in itens
    ])


@bp.post("/api/planejamento")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_create():
    data = request.get_json()
    exists = Planejamento.query.filter_by(
        data=data["data"],
        turno=data["turno"],
        instrutor_id=data["instrutor_id"],
    ).first()
    if exists:
        return jsonify({"error": "Conflito de agenda"}), 409
    p = Planejamento(
        data=data["data"],
        turno=data["turno"],
        carga_horas=data.get("carga_horas"),
        modalidade=data.get("modalidade"),
        treinamento=data["treinamento"],
        instrutor_id=data["instrutor_id"],
        local=data.get("local"),
        cliente=data.get("cliente"),
        observacao=data.get("observacao"),
        status=data.get("status", "Planejado"),
        origem=data.get("origem", "Manual"),
    )
    db.session.add(p)
    db.session.commit()
    return jsonify({"id": p.id}), 201


@bp.put("/api/planejamento/<int:pid>")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_update(pid):
    p = Planejamento.query.get_or_404(pid)
    data = request.get_json()
    if (
        "data" in data
        or "turno" in data
        or "instrutor_id" in data
    ):
        new_data = data.get("data", p.data)
        new_turno = data.get("turno", p.turno)
        new_instrutor = data.get("instrutor_id", p.instrutor_id)
        exists = Planejamento.query.filter_by(
            data=new_data,
            turno=new_turno,
            instrutor_id=new_instrutor,
        ).first()
        if exists and exists.id != p.id:
            return jsonify({"error": "Conflito de agenda"}), 409
    for f in [
        "data",
        "turno",
        "carga_horas",
        "modalidade",
        "treinamento",
        "instrutor_id",
        "local",
        "cliente",
        "observacao",
        "status",
    ]:
        if f in data:
            setattr(p, f, data[f])
    db.session.commit()
    return jsonify({"ok": True})


@bp.delete("/api/planejamento/<int:pid>")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_delete(pid):
    p = Planejamento.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({"ok": True})


@bp.post("/api/planejamento/import")
@require_roles(ROLE_ADMIN, ROLE_GESTOR)
def api_import():
    f = request.files["file"]
    wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
    ws = wb.active
    headers = [
        str((ws.cell(row=1, column=i).value or "")).strip().upper()
        for i in range(1, ws.max_column + 1)
    ]

    def col(name):
        return headers.index(name) + 1 if name in headers else None

    if col("DATA") and col("TURNO") and col("TREINAMENTO") and col("INSTRUTOR"):
        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, col("DATA")).value
            if not d:
                continue
            if hasattr(d, "date"):
                data_val = d.date()
            else:
                data_val = datetime.strptime(str(d), "%d/%m/%Y").date()
            turno = (
                (ws.cell(r, col("TURNO")).value or "")
                .upper()
                .replace("Ã", "A")
                .replace("Â", "A")
            )
            instrutor_nome = str(ws.cell(r, col("INSTRUTOR")).value or "").strip()
            instrutor_id = _get_instrutor_id(instrutor_nome)
            p = Planejamento(
                data=data_val,
                turno=turno,
                carga_horas=ws.cell(r, col("CARGA")).value if col("CARGA") else None,
                modalidade=ws.cell(r, col("MODALIDADE")).value if col("MODALIDADE") else None,
                treinamento=str(ws.cell(r, col("TREINAMENTO")).value or ""),
                instrutor_id=instrutor_id,
                local=ws.cell(r, col("LOCAL")).value if col("LOCAL") else None,
                cliente=ws.cell(r, col("CLIENTE")).value if col("CLIENTE") else None,
                observacao=ws.cell(r, col("OBS")).value if col("OBS") else None,
            )
            db.session.add(p)
        db.session.commit()
        return jsonify({"ok": True})
    abort(400, "Formato de planilha não reconhecido")


@bp.get("/api/planejamento/export")
@require_roles(ROLE_ADMIN, ROLE_GESTOR, ROLE_USER)
def api_export():
    mes = request.args.get("mes")
    y, m = [int(x) for x in mes.split("-")]
    q = Planejamento.query.filter(
        extract("year", Planejamento.data) == y,
        extract("month", Planejamento.data) == m,
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Planejamento"
    ws.append(
        [
            "DATA",
            "SEMANA",
            "TURNO",
            "CARGA",
            "MODALIDADE",
            "TREINAMENTO",
            "INSTRUTOR",
            "LOCAL",
            "CLIENTE",
            "STATUS",
            "OBSERVACAO",
        ]
    )
    for p in q:
        d = p.data
        ws.append(
            [
                d.strftime("%d/%m/%Y"),
                ["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sab"][d.weekday()],
                p.turno,
                p.carga_horas,
                p.modalidade,
                p.treinamento,
                p.instrutor.nome if p.instrutor else "",
                p.local,
                p.cliente,
                p.status,
                p.observacao or "",
            ]
        )
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"planejamento_{mes}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
