import io
from datetime import date

from openpyxl import Workbook, load_workbook

from src.models import db, Planejamento
from src.models.instrutor import Instrutor


def test_api_planejamento_export(client, non_admin_auth_headers):
    with client.application.app_context():
        instr = Instrutor(nome="Instrutor Teste")
        db.session.add(instr)
        db.session.flush()
        p = Planejamento(
            data=date(2024, 1, 5),
            turno="MANHA",
            carga_horas=8,
            modalidade="Presencial",
            treinamento="Curso X",
            instrutor_id=instr.id,
            local="Sala 1",
            cliente="Cliente",
            status="Planejado",
            origem="Manual",
        )
        db.session.add(p)
        db.session.commit()
    resp = client.get(
        "/planejamento/api/planejamento/export?mes=2024-01",
        headers=non_admin_auth_headers,
    )
    assert resp.status_code == 200
    wb = load_workbook(filename=io.BytesIO(resp.data))
    ws = wb.active
    assert ws.title == "Planejamento"
    assert ws.cell(row=2, column=1).value == "05/01/2024"
    assert ws.cell(row=2, column=4).value == "Instrutor Teste"

def test_api_planejamento_import(client, login_admin):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "DATA",
        "SEMANA",
        "TURNO",
        "CARGA",
        "MODALIDADE",
        "TREINAMENTO",
        "INSTRUTOR",
        "LOCAL",
        "CLIENTE",
        "OBS",
    ])
    ws.append([
        "05/01/2024",
        "Sex",
        "MANHA",
        8,
        "Presencial",
        "Curso Y",
        "Instrutor Novo",
        "Sala",
        "Cliente",
        "Observacao",
    ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    data = {"file": (stream, "import.xlsx")}
    token, _ = login_admin(client)
    headers = {"Authorization": f"Bearer {token}"}
    resp = client.post(
        "/planejamento/api/planejamento/import",
        data=data,
        content_type="multipart/form-data",
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.json["ok"] is True
    with client.application.app_context():
        itens = Planejamento.query.all()
        assert len(itens) == 1
        p = itens[0]
        assert p.treinamento == "Curso Y"
        instrutor = Instrutor.query.get(p.instrutor_id)
        assert instrutor.nome == "Instrutor Novo"


def test_api_planejamento_conflict(client, login_admin):
    with client.application.app_context():
        instr = Instrutor(nome="Instrutor Conflito")
        db.session.add(instr)
        db.session.flush()
        instr_id = instr.id
        p1 = Planejamento(
            data=date(2024, 1, 1),
            turno="MANHA",
            treinamento="Curso 1",
            instrutor_id=instr_id,
        )
        p2 = Planejamento(
            data=date(2024, 1, 2),
            turno="TARDE",
            treinamento="Curso 2",
            instrutor_id=instr_id,
        )
        db.session.add_all([p1, p2])
        db.session.commit()
        pid2 = p2.id
    token, _ = login_admin(client)
    headers = {"Authorization": f"Bearer {token}"}
    resp = client.post(
        "/planejamento/api/planejamento",
        json={
            "data": "2024-01-01",
            "turno": "MANHA",
            "treinamento": "Curso X",
            "instrutor_id": instr_id,
        },
        headers=headers,
    )
    assert resp.status_code == 409
    resp2 = client.put(
        f"/planejamento/api/planejamento/{pid2}",
        json={"data": "2024-01-01", "turno": "MANHA"},
        headers=headers,
    )
    assert resp2.status_code == 409
